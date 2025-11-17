# web/views.py
from datetime import datetime
import json
from django.db import connection
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import joblib
import numpy as np
import pandas as pd
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from pathlib import Path
from math import log1p
from .models import ReportDownloadHistory
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.http import HttpResponse
from io import BytesIO
import json
import traceback
from datetime import datetime
from decimal import Decimal
from django.http import JsonResponse, HttpResponse
from .models import ReportDownloadHistory
from django.utils import timezone
from datetime import timedelta


def login_view(request):
    # Nếu user đã đăng nhập, chuyển đến home
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')  # hoặc 'email' nếu dùng email
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_superuser:  # chỉ cho phép superuser đăng nhập
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, "Tài khoản này không có quyền truy cập.")
        else:
            messages.error(request, "Sai tài khoản hoặc mật khẩu.")

    return render(request, 'web/login.html')


@login_required(login_url='login')
def home(request):
    return redirect('overview')


def logout_view(request):
    logout(request)
    return redirect('login')


# Load model
BASE_DIR = Path(__file__).resolve().parent.parent
MODEL_PATH = BASE_DIR / "web" / "ml_model" / "lightgbm_v8_grid.pkl"
model = joblib.load(MODEL_PATH)


def q(sql):
    with connection.cursor() as cur:
        cur.execute("SET NAMES utf8mb4;")
        cur.execute(sql)
        return cur.fetchall()


# PAGE
@login_required(login_url='login')
def price_view(request):
    rows = q("SELECT variable_name, original_value, encoded_value FROM mapping")
    mapping = {}
    for var, original, encoded in rows:
        mapping.setdefault(var, []).append({"id": int(encoded), "name": str(original)})

    ctx = {
        "product_groups": mapping.get("product_group", []),
        "brands": mapping.get("brand_name", []),
        "regions": mapping.get("region", []),
    }
    return render(request, "web/price.html", ctx)


from datetime import datetime, timedelta, date


@login_required(login_url='login')
def price_view(request):
    rows = q("SELECT variable_name, original_value, encoded_value FROM mapping")
    mapping = {}
    for var, original, encoded in rows:
        mapping.setdefault(var, []).append({"id": int(encoded), "name": str(original)})

    # Lấy thông tin thời gian hiện tại
    today = datetime.today()
    year, week_num, _ = today.isocalendar()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    ctx = {
        "product_groups": mapping.get("product_group", []),
        "brands": mapping.get("brand_name", []),
        "regions": mapping.get("region", []),
        "today": today.strftime("%d/%m/%Y"),
        "week_num": week_num,
        "week_range": f"{start_of_week.strftime('%d/%m')} – {end_of_week.strftime('%d/%m/%Y')}",
        "year": year,
    }
    return render(request, "web/price.html", ctx)


FEATURES_V8 = [
    # Region (7)
    "region_KVCA", "region_KVMB", "region_KVMT", "region_KVMTR", "region_KVTN", "region_KVMN", "region_Khac",
    # Urbanization (4)
    "urbanization_Noi_thanh", "urbanization_Nong_thon", "urbanization_TT_hanh_chinh_kinh_te", "urbanization_Khac",
    # Encoded (3)
    "product_group_enc", "brand_name_enc", "price_group_enc",
    # Numeric & Time (8)
    "net_price", "discount_rate", "margin", "avg_weekly_sales", "promo_flag", "price_relative",
    "holiday_flag_soft", "week_num"
]

def _predict_internal_v8(payload):
    region_enc = int(payload.get("region_enc", 0))  # 0..6
    product_group_enc = int(payload.get("product_group_enc", 0))
    brand_name_enc = int(payload.get("brand_name_enc", 0))
    net_price = float(payload.get("net_price", 4_000_000))
    discount_pct_ui = float(payload.get("discount_rate", 20.0))  # % từ UI
    promo_on = bool(payload.get("promo_on", True))

    cost_price = 0.70 * net_price
    margin = max(net_price - cost_price, 0.0)
    promo_flag = 1 if promo_on else 0
    discount_pct_ui = discount_pct_ui if promo_on else 0.0

    # price_relative
    baseline_price = 3_000_000
    price_relative = (net_price / baseline_price) if baseline_price > 0 else 1.0

    # Không có chuỗi lịch sử và lịch lễ theo SKU → baseline nhỏ
    avg_weekly_sales = 0.2
    holiday_flag_soft = 0.1

    # Tuần hiện tại
    week_num = date.today().isocalendar().week
    discount_rate_n = max(0.0, min(discount_pct_ui, 50.0)) / 50.0
    net_price_n = min(net_price / 10_000_000.0, 1.0)
    margin_n = min(margin / 10_000_000.0, 1.0)
    price_rel_n = min(price_relative / 2.0, 1.0)

    region_ohe = [0] * 7
    if 0 <= region_enc < 7:
        region_ohe[region_enc] = 1
    urban_ohe = [0, 0, 0, 0]

    # price_group_enc gộp theo net_price
    price_group_enc = int(payload.get(
        "price_group_enc",
        1 if net_price < 2_000_000 else (2 if net_price < 5_000_000 else 3)
    ))

    row = {
        # region 7
        "region_KVCA": region_ohe[0], "region_KVMB": region_ohe[1], "region_KVMT": region_ohe[2],
        "region_KVMTR": region_ohe[3], "region_KVTN": region_ohe[4], "region_KVMN": region_ohe[5],
        "region_Khac": region_ohe[6],
        # urbanization 4
        "urbanization_Noi_thanh": urban_ohe[0], "urbanization_Nong_thon": urban_ohe[1],
        "urbanization_TT_hanh_chinh_kinh_te": urban_ohe[2], "urbanization_Khac": urban_ohe[3],
        # encodes 3
        "product_group_enc": product_group_enc,
        "brand_name_enc": brand_name_enc,
        "price_group_enc": price_group_enc,
        # numeric & time 8
        "net_price": net_price_n,
        "discount_rate": discount_rate_n,
        "margin": margin_n,
        "avg_weekly_sales": avg_weekly_sales,
        "promo_flag": promo_flag,
        "price_relative": price_rel_n,
        "holiday_flag_soft": holiday_flag_soft,
        "week_num": week_num,
    }

    X = pd.DataFrame([row], columns=FEATURES_V8)
    y_pred = float(model.predict(X)[0])
    sold = max(y_pred, 0.0)
    revenue = sold * net_price
    profit = sold * margin
    return {"sold": sold, "revenue": revenue, "profit": profit}


@csrf_exempt
@login_required(login_url='login')
def predict(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        y = _predict_internal_v8(payload)

        net_price = float(payload.get("net_price", 4_000_000))
        # discount %
        discount_pct_ui = float(payload.get("discount_rate", 20.0))
        promo_on = bool(payload.get("promo_on", True))
        if not promo_on:
            discount_pct_ui = 0.0

        # biên lợi nhuận %
        cost_price = 0.7 * net_price
        margin = max(net_price - cost_price, 0.0)
        margin_rate = (margin / net_price * 100) if net_price else 0.0

        return JsonResponse({
            "ok": True,
            "predicted_sales": round(y["sold"], 2),
            "predicted_revenue": round(y["revenue"], 2),
            "predicted_profit": round(y["profit"], 2),
            "margin_rate": round(margin_rate, 2),
        })
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


#SIMULATE SERIES (cho biểu đồ)
@csrf_exempt
@login_required(login_url='login')
def simulate_series(request):
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        mode = payload.get("mode", "discount")

        # defaults an toàn
        payload.setdefault("region_enc", 0)
        payload.setdefault("product_group_enc", 0)
        payload.setdefault("brand_name_enc", 0)
        payload.setdefault("net_price", 4_000_000.0)
        payload.setdefault("discount_rate", 20.0)
        payload.setdefault("promo_on", True)

        series = []
        base_price = float(payload["net_price"])

        if mode == "price":
            # quét +-30% giá theo bước 5%
            for i in range(-6, 7):
                p = base_price * (1 + i * 0.05)
                temp = {**payload, "net_price": p}
                y = _predict_internal_v8(temp)
                series.append({"x": round(p), **y})

        elif mode == "discount":
            # quét 0..50% discount theo bước 5%
            for d in range(0, 55, 5):
                temp = {**payload, "discount_rate": d, "promo_on": True}
                y = _predict_internal_v8(temp)
                series.append({"x": d, **y})

        elif mode == "region":
            # quét 7 vùng
            for r in range(7):
                temp = {**payload, "region_enc": r}
                y = _predict_internal_v8(temp)
                series.append({"x": r, **y})

        return JsonResponse({"ok": True, "series": series})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})


from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required


def create_excel_content(payload, wb):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference

    # ===== DEFAULT =====
    payload.setdefault("region_enc", 0)
    payload.setdefault("product_group_enc", 0)
    payload.setdefault("brand_name_enc", 0)
    payload.setdefault("net_price", 4_000_000.0)
    payload.setdefault("discount_rate", 20.0)
    payload.setdefault("promo_on", True)

    # ===== VÙNG MAPPING =====
    region_names = ["KVCA", "KVMB", "KVMT", "KVMTR", "KVTN", "KVMN", "Khác"]

    # ===== TẠO SERIES =====
    series_price = []
    series_discount = []
    series_region = []

    base_price = float(payload["net_price"])

    # --- PRICE ±30% ---
    for i in range(-6, 7):
        p = base_price * (1 + i * 0.05)
        y = _predict_internal_v8({**payload, "net_price": p})
        series_price.append({"x": p, **y})

    # --- DISCOUNT 0..50% ---
    for d in range(0, 55, 5):
        y = _predict_internal_v8({**payload, "discount_rate": d, "promo_on": True})
        series_discount.append({"x": d, **y})

    # --- REGION 0..6 ---
    for r in range(7):
        y = _predict_internal_v8({**payload, "region_enc": r})
        series_region.append({"x": r, **y})

    # ===== OPTIMIZATION =====
    best_price = max(series_price, key=lambda x: x["profit"])

    # discount: ưu tiên lợi nhuận > nếu bằng nhau → ưu tiên discount thấp
    best_discount = series_discount[0]
    for cur in series_discount:
        if cur["profit"] > best_discount["profit"]:
            best_discount = cur
        elif cur["profit"] == best_discount["profit"] and cur["x"] < best_discount["x"]:
            best_discount = cur

    best_region = max(series_region, key=lambda x: x["profit"])

    # ===== BASE STYLES =====
    bold = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="C5D9F1")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    money_fmt = "#,##0₫"

    # ===============================
    # 1️⃣ SUMMARY SHEET
    # ===============================
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 30

    ws_sum.append(["HẠNG MỤC", "GIÁ TRỊ"])
    for c in ws_sum[1]:
        c.font = bold
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    summary_rows = [
        ["Giá bán tối ưu (VND)", round(best_price["x"])],
        ["Khuyến mãi tối ưu (%)", best_discount["x"]],
        ["Vùng lợi nhuận cao nhất", region_names[best_region["x"]]],
        ["Lợi nhuận tối đa (VND)", best_discount["profit"]],
        ["Doanh thu tối đa (VND)", best_discount["revenue"]],
        ["Sản lượng tối đa (sp/tuần)", best_discount["sold"]]
    ]

    for label, value in summary_rows:
        ws_sum.append([label, value])
        row_idx = ws_sum.max_row

        # style
        ws_sum[f"A{row_idx}"].border = border
        ws_sum[f"B{row_idx}"].border = border

        # FORMAT THEO LOẠI
        if "VND" in label:
            ws_sum[f"B{row_idx}"].number_format = money_fmt
        elif "(%)" in label:
            ws_sum[f"B{row_idx}"].value = value / 100  # Excel % format = 0.05
            ws_sum[f"B{row_idx}"].number_format = "0%"

    # GỢI Ý AI
    ws_sum.append([])
    ws_sum.append([
        "Gợi ý AI",
        f"Áp dụng giá {round(best_price['x']):,}đ, giảm {best_discount['x']}% tại "
        f"{region_names[best_region['x']]} đạt lợi nhuận {best_discount['profit']:,}đ/tuần."
    ])
    ws_sum["A" + str(ws_sum.max_row)].font = Font(bold=True, color="0070C0")
    ws_sum["B" + str(ws_sum.max_row)].font = Font(bold=True, color="0070C0")
    ws_sum["B" + str(ws_sum.max_row)].alignment = Alignment(wrap_text=True)

    # ===============================
    # 2️⃣ PRICE SIMULATION SHEET
    # ===============================
    ws1 = wb.create_sheet("Price Simulation")
    ws1.append(["Giá bán (VND)", "Sản lượng", "Doanh thu (VND)", "Lợi nhuận (VND)"])
    for c in ws1[1]:
        c.font = bold
        c.fill = header_fill
        c.border = border

    for r in series_price:
        ws1.append([round(r["x"]), r["sold"], r["revenue"], r["profit"]])
        for c in ws1[ws1.max_row]:
            c.border = border
            if c.col_idx >= 3:
                c.number_format = money_fmt

    chart = LineChart()
    chart.title = "Biến động theo Giá bán"
    chart.y_axis.title = "Giá trị (VND)"
    chart.x_axis.title = "Giá bán (VND)"
    data = Reference(ws1, min_col=2, max_col=4, min_row=1, max_row=ws1.max_row)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws1.add_chart(chart, "G3")

    # ===============================
    # 3️⃣ DISCOUNT SIMULATION
    # ===============================
    ws2 = wb.create_sheet("Discount Simulation")
    ws2.append(["Giảm giá (%)", "Sản lượng", "Doanh thu (VND)", "Lợi nhuận (VND)"])
    for c in ws2[1]:
        c.font = bold
        c.fill = header_fill
        c.border = border

    for r in series_discount:
        ws2.append([r["x"], r["sold"], r["revenue"], r["profit"]])
        for c in ws2[ws2.max_row]:
            c.border = border
            if c.col_idx >= 3:
                c.number_format = money_fmt

    chart2 = LineChart()
    chart2.title = "Tác động của Mức khuyến mãi"
    chart2.y_axis.title = "Giá trị (VND)"
    chart2.x_axis.title = "Mức giảm giá (%)"
    data = Reference(ws2, min_col=2, max_col=4, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws2.add_chart(chart2, "G3")

    # ===============================
    # 4️⃣ REGION SIMULATION
    # ===============================
    ws3 = wb.create_sheet("Region Simulation")
    ws3.append(["Vùng", "Sản lượng", "Doanh thu (VND)", "Lợi nhuận (VND)"])
    for c in ws3[1]:
        c.font = bold
        c.fill = header_fill
        c.border = border

    for r in series_region:
        label = region_names[r["x"]] if 0 <= r["x"] < len(region_names) else "?"
        ws3.append([label, r["sold"], r["revenue"], r["profit"]])
        for c in ws3[ws3.max_row]:
            c.border = border
            if c.col_idx >= 3:
                c.number_format = money_fmt

    chart3 = BarChart()
    chart3.title = "So sánh Sản lượng theo Vùng"
    chart3.x_axis.title = "Vùng"
    chart3.y_axis.title = "Sản lượng"
    data = Reference(ws3, min_col=2, max_col=2, min_row=1, max_row=ws3.max_row)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    ws3.add_chart(chart3, "G3")


@login_required(login_url='login')
def forecast_view(request):
    """
    Trang Dự báo nhu cầu: lấy danh sách sản phẩm, khu vực, và TÊN GỐC CÁC MÔ HÌNH.
    """
    # ====== Lấy danh sách sản phẩm ======
    try:
        with connection.cursor() as cur:
            cur.execute("SET NAMES utf8mb4;")
            cur.execute("""
                SELECT DISTINCT product_group
                FROM product
                WHERE product_group IS NOT NULL
                ORDER BY product_group
            """)
            products = [row[0] for row in cur.fetchall()]
    except Exception as e:
        print(f"Lỗi truy vấn sản phẩm: {e}")
        products = []

    # ====== Lấy danh sách khu vực ======
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT original_value
                FROM mapping
                WHERE variable_name = 'region'
                ORDER BY encoded_value
            """)
            regions = [row[0] for row in cur.fetchall()]
    except Exception as e:
        print(f"Lỗi truy vấn khu vực: {e}")
        regions = []

    # Tim mo hinh
    DEFAULT_MODEL = 'lightgbm_v8_grid'

    try:
        # Giả định: thư mục 'ml_model' nằm ngang hàng với thư mục 'web' chứa views.py
        BASE_DIR = Path(__file__).resolve().parent
        MODEL_DIR = BASE_DIR.parent / 'ml_model'

        # Lấy danh sách tên gốc (stem) của file, không có .pkl
        model_stems = [p.stem for p in MODEL_DIR.glob('*.pkl') if not p.name.startswith('.')]

        # Xử lý trường hợp thư mục rỗng
        if not model_stems:
            model_stems = [DEFAULT_MODEL]

    except Exception as e:
        print(f"Lỗi khi quét thư mục mô hình: {e}")
        model_stems = [DEFAULT_MODEL]

    ctx = {
        "products": products,
        "regions": regions,
        "models": model_stems  # Danh sách tên gốc (VD: ['ligth_gbm_v8_grid', 'model_b'])
    }
    return render(request, "web/forecast.html", ctx)

def _fetch_sales_weekly(product_code: str, aggregate_by_month=True):
    if product_code.upper() == "ALL":
        base_sql = """
            SELECT week, SUM(sold_quantity) AS qty,
                   AVG(lag_1), AVG(rolling_mean_4w), AVG(growth_rate)
            FROM sales_weekly
            GROUP BY week
            ORDER BY week
        """
        with connection.cursor() as cur:
            cur.execute("SET NAMES utf8mb4;")
            cur.execute(base_sql)
            rows = cur.fetchall()
        if not rows:
            return {"months": [], "actual": [], "lag1": [], "roll4": [], "growth": []}
        rows = [(int(w), float(q or 0), float(l1 or 0), float(r4 or 0), float(g or 0)) for (w, q, l1, r4, g) in rows]
    else:
        # Giữ nguyên các phần xử lý cho DEP, product_id,...
        if len(product_code) <= 5:
            base_sql = """
                SELECT s.week, SUM(s.sold_quantity) AS qty,
                    AVG(s.lag_1), AVG(s.rolling_mean_4w), AVG(s.growth_rate)
                FROM sales_weekly s
                JOIN product p ON s.product_id = p.product_id
                WHERE p.product_group = %s
                GROUP BY s.week
                ORDER BY s.week
            """
        else:
            base_sql = """
                SELECT week, sold_quantity AS qty, lag_1, rolling_mean_4w, growth_rate
                FROM sales_weekly
                WHERE product_id = %s
                ORDER BY week
            """
        with connection.cursor() as cur:
            cur.execute("SET NAMES utf8mb4;")
            cur.execute(base_sql, [product_code])
            rows = cur.fetchall()
            rows = [(int(w), float(q or 0), float(l1 or 0), float(r4 or 0), float(g or 0)) for (w, q, l1, r4, g) in
                    rows]

    # Gộp dữ liệu theo tháng
    if aggregate_by_month:
        months, actual_m, lag1_m, roll4_m, growth_m = [], [], [], [], []
        month_idx, temp = 1, []

        for i, (w, q, l1, r4, g) in enumerate(rows, start=1):
            temp.append((q, l1, r4, g))
            if i % 4 == 0 or i == len(rows):
                avg_q = sum(x[0] for x in temp) / len(temp)
                avg_l1 = sum(x[1] for x in temp) / len(temp)
                avg_r4 = sum(x[2] for x in temp) / len(temp)
                avg_g = sum(x[3] for x in temp) / len(temp)

                months.append(month_idx)
                actual_m.append(round(avg_q, 2))
                lag1_m.append(round(avg_l1, 2))
                roll4_m.append(round(avg_r4, 2))
                growth_m.append(round(avg_g, 2))

                temp.clear()
                month_idx += 1

        return {
            "months": months,
            "actual": actual_m,
            "lag1": lag1_m,
            "roll4": roll4_m,
            "growth": growth_m,
        }

    # Nếu không gộp, trả về theo tuần
    weeks, actual, lag1, roll4, growth = [], [], [], [], []
    for w, q, l1, r4, g in rows:
        weeks.append(int(w))
        actual.append(float(q))
        lag1.append(float(l1))
        roll4.append(float(r4))
        growth.append(float(g))

    return {"weeks": weeks, "actual": actual, "lag1": lag1, "roll4": roll4, "growth": growth}


# SMA Baseline
def _sma_forecast(actual, lag1=None, roll4=None, horizon=6, alpha=0.6):
    n = len(actual)
    fitted = []
    for i in range(n):
        prev = actual[i - 1] if i > 0 else actual[0]
        window = actual[max(0, i - 3):i + 1]
        roll = sum(window) / len(window) if len(window) else prev
        pred = alpha * prev + (1 - alpha) * roll
        fitted.append(round(pred, 2))

    # Dự báo tương lai
    future = []
    last = fitted[-1] if fitted else actual[-1]
    last_roll = sum(actual[-4:]) / 4 if len(actual) >= 4 else actual[-1]
    for _ in range(horizon):
        pred = alpha * last + (1 - alpha) * last_roll
        pred = max(pred, 0.0)
        future.append(round(pred, 2))
        last = pred
        last_roll = (last_roll * 3 + pred) / 4
    return fitted, future


def _mape(y_true, y_pred):
    eps = 1e-9
    n = min(len(y_true), len(y_pred))
    if n == 0:
        return None
    s = 0.0
    c = 0
    for i in range(n):
        if y_true[i] == 0:
            continue
        s += abs(y_true[i] - y_pred[i]) / max(abs(y_true[i]), eps)
        c += 1
    return round(100.0 * s / c, 2) if c > 0 else None

# API
@csrf_exempt
@require_http_methods(["POST"])
@login_required(login_url='login')
def api_forecast(request):

    try:
        # ======================================================
        # 1️⃣ Nhận payload từ frontend
        # ======================================================
        raw = (request.body or b"").decode("utf-8").strip()
        payload = json.loads(raw) if raw else {}

        horizon = int(payload.get("horizon", 3))
        product_id = (payload.get("product_id") or "ALL").strip().upper()
        region_filter = (payload.get("region") or "ALL").strip().upper()

        # ======================================================
        # 2️⃣ Lấy dữ liệu bán hàng (tháng) – nhóm hoặc sản phẩm
        # ======================================================
        series = _fetch_sales_weekly(product_id, aggregate_by_month=True)
        months = series.get("months") or series.get("weeks") or []
        actual = series.get("actual", [])
        lag1 = series.get("lag1", [])
        roll4 = series.get("roll4", [])

        if not months or not actual:
            return JsonResponse({
                "ok": False,
                "error": f"Không có dữ liệu cho product_id={product_id}"
            }, status=404)

        # ======================================================
        # 3️⃣ Dự báo bằng mô hình SMA baseline
        # ======================================================
        fitted, future = _sma_forecast(actual, lag1, roll4, horizon=horizon, alpha=0.6)

        # Tạo nhãn tương lai (theo tháng)
        last_m = months[-1]
        future_months = [last_m + i for i in range(1, horizon + 1)]

        # KPI: tổng và trung bình dự báo
        hist_tail = actual[-horizon:] if len(actual) >= horizon else actual
        fit_tail = fitted[-len(hist_tail):]
        mape = _mape(hist_tail, fit_tail)
        sum_future = round(float(sum(future)), 2)
        avg_future = round(sum_future / max(horizon, 1), 2)
        kpis = {"sum_forecast": sum_future, "avg_forecast": avg_future, "mape_tail": mape}

        # ======================================================
        # 4️⃣ Cơ cấu nhu cầu theo khu vực
        # ======================================================
        with connection.cursor() as cur:
            sql_region = """
                SELECT 
                    m.original_value AS region, 
                    SUM(s.sold_quantity) AS total_qty
                FROM sales s
                INNER JOIN product p ON s.product_id = p.product_id
                LEFT JOIN mapping m 
                    ON m.encoded_value = s.distribution_channel_code_enc
                   AND m.variable_name = 'region'
                WHERE (%s = 'ALL' OR p.product_group = %s)
                GROUP BY m.original_value
                HAVING total_qty > 0
                ORDER BY total_qty DESC;
            """
            cur.execute(sql_region, [product_id, product_id])
            region_rows = cur.fetchall()

        region_labels = [r[0] for r in region_rows]
        region_values = [float(r[1]) for r in region_rows]

        # ======================================================
        # 5️⃣ Top sản phẩm tăng/giảm trong nhóm đang chọn
        # ======================================================
        with connection.cursor() as cur:
            sql_top = """
                WITH avg_group AS (
                    SELECT p2.product_group, AVG(s2.sold_quantity) AS group_avg
                    FROM sales s2
                    JOIN product p2 ON s2.product_id = p2.product_id
                    GROUP BY p2.product_group
                )
                SELECT 
                    p.product_id,
                    p.product_group,
                    ROUND(AVG(s.sold_quantity), 2) AS avg_sales,
                    ROUND(
                        ((AVG(s.sold_quantity) - ag.group_avg) / NULLIF(ag.group_avg, 0)) * 100,
                        2
                    ) AS pct_change
                FROM sales s
                JOIN product p ON s.product_id = p.product_id
                JOIN avg_group ag ON p.product_group = ag.product_group
                WHERE (%s = 'ALL' OR p.product_group = %s)
                GROUP BY p.product_group, p.product_id
                ORDER BY pct_change DESC
                LIMIT 10;
            """
            cur.execute(sql_top, [product_id, product_id])
            rows_top = cur.fetchall()

        top_labels = [r[0] for r in rows_top]
        top_changes = [round(float(r[3] or 0), 2) for r in rows_top]

        top_strongest = top_labels[0] if top_labels else None
        top_strongest_change = top_changes[0] if top_changes else None
        trend = "tăng" if (top_changes and top_changes[0] > 0) else "giảm"

        # ======================================================
        # 6️⃣ Trả JSON đầy đủ cho frontend
        # ======================================================
        return JsonResponse({
            "ok": True,
            "product_id": product_id,
            "labels_hist": months,
            "actual": actual,
            "fitted": [round(x, 2) for x in fitted],
            "labels_future": future_months,
            "forecast": [round(x, 2) for x in future],
            "kpis": kpis,
            "region_labels": region_labels,
            "region_data": region_values,
            "top_labels": top_labels,
            "top_changes": top_changes,
            "top_strongest": top_strongest,
            "top_strongest_change": top_strongest_change,
            "trend": trend,
        })

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

@login_required(login_url='login')
def report_view(request):
    """
    Hàm này sẽ render template report.html.
    Template này nằm trong web/templates/web/report.html
    """
    context = {} # Bạn có thể thêm dữ liệu để truyền cho template ở đây
    return render(request, 'web/report.html', context)

@api_view(['GET'])
def api_reports(request):
    # Query params
    search = request.GET.get('search', '').lower()
    report_type = request.GET.get('type', 'all')
    creator = request.GET.get('creator', '').lower()
    time_filter = request.GET.get('time', 'all')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('pageSize', 5))

    reports = ReportDownloadHistory.objects.all().select_related('user')

    # ============================
    # 🔍 SEARCH (KHÔNG DÙNG Q)
    # ============================
    if search:
        qs1 = ReportDownloadHistory.objects.filter(filename__icontains=search)
        qs2 = ReportDownloadHistory.objects.filter(user__username__icontains=search)
        reports = qs1.union(qs2)  # OR logic

    # ============================
    # 📌 Filter loại báo cáo
    # ============================
    if report_type != 'all':
        reports = reports.filter(payload__type=report_type)

    # ============================
    # 👤 Filter theo người tạo
    # ============================
    if creator:
        reports = reports.filter(user__username__icontains=creator)

    # ============================
    # ⏳ Time filter
    # ============================
    from datetime import datetime, timedelta
    now = datetime.now()

    if time_filter != 'all':
        delta = {
            'today': 0,
            '3_days': 3,
            '7_days': 7,
            '30_days': 30,
            '90_days': 90,
            '1_year': 365
        }.get(time_filter, 0)

        if delta == 0:
            reports = reports.filter(created_at__date=now.date())
        else:
            reports = reports.filter(created_at__gte=now - timedelta(days=delta))

    # ============================
    # 📄 Pagination
    # ============================
    total = reports.count()
    total_pages = (total + page_size - 1) // page_size
    start = (page - 1) * page_size
    end = start + page_size

    reports = reports.order_by('-created_at')[start:end]

    # ============================
    # 📤 Response
    # ============================
    data = []
    for r in reports:
        data.append({
            'id': r.id,
            'name': r.filename,
            'type': r.payload.get('type', 'Khác'),
            'creator': r.user.username,
            'date': r.created_at.strftime('%d/%m/%Y'),
            'size': f"{r.file_size_kb:.2f} KB"
        })

    return Response({
        'reports': data,
        'currentPage': page,
        'totalPages': total_pages
    })


def download_history_view(request):
    queryset = ReportDownloadHistory.objects.all().order_by('-created_at')
    time_filter = request.GET.get("time", "all")

    if time_filter != "all":
        now = timezone.now()

        if time_filter == "today":
            queryset = queryset.filter(created_at__date=now.date())

        elif time_filter == "7days":
            queryset = queryset.filter(created_at__gte=now - timedelta(days=7))

        elif time_filter == "30days":
            queryset = queryset.filter(created_at__gte=now - timedelta(days=30))

    # ĐỔ RA TEMPLATE DƯỚI TÊN "history"
    return render(request, 'reports/history.html', {
        'history': queryset
    })
@api_view(['DELETE'])
@login_required(login_url='login')
def api_delete_report(request, report_id):
    try:
        report = ReportDownloadHistory.objects.get(id=report_id)
        report.delete()
        return Response({"ok": True}, status=204)

    except ReportDownloadHistory.DoesNotExist:
        return Response({"ok": False,"error": "Không tìm thấy báo cáo ID này."}, status=404)

    except Exception as e:
        return Response({"ok": False,"error": f"Lỗi server không xác định: {str(e)}"}, status=500)

@api_view(['GET'])
@login_required(login_url='login')
def api_redownload_report(request, report_id):
    try:
        report = ReportDownloadHistory.objects.get(id=report_id, user=request.user)

        # Lấy tham số mô phỏng gốc đã lưu trong DB
        params = report.payload.get('params', {})

        wb = openpyxl.Workbook()

        create_excel_content(params, wb)

        # Lưu Workbook vào stream
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Trả về response, dùng tên file đã lưu trong DB
        response = HttpResponse(
            file_stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{report.filename}"'

        return response

    except ReportDownloadHistory.DoesNotExist:
        return Response({"error": "Không tìm thấy báo cáo hoặc bạn không có quyền truy cập."}, status=404)
    except Exception as e:
        return Response({"error": f"Lỗi tái tạo báo cáo: {str(e)}"}, status=500)


    # web/views.py
# ==============================================================
# Helper
# ==============================================================
def _to_float(x):
    if x is None:
        return 0.0
    if isinstance(x, Decimal):
        return float(x)
    try:
        return float(x)
    except Exception:
        return 0.0


# ==============================================================
# 1. RENDER OVERVIEW PAGE (dropdowns)
# ==============================================================
@login_required(login_url='login')
def overview_view(request):
    with connection.cursor() as cur:
        cur.execute("SET NAMES utf8mb4;")
        cur.execute("""
            SELECT DISTINCT product_group
            FROM product
            WHERE product_group IS NOT NULL
            ORDER BY product_group
        """)
        products = [row[0] for row in cur.fetchall()]

    with connection.cursor() as cur:
        cur.execute("""
            SELECT original_value
            FROM mapping
            WHERE variable_name = 'region'
            ORDER BY encoded_value
        """)
        regions = [row[0] for row in cur.fetchall()]

    ctx = {
        "products": products,
        "regions": regions
    }
    return render(request, "web/overview.html", ctx)


# 2. API OVERVIEW
@csrf_exempt
@require_http_methods(["GET"])
@login_required(login_url='login')
def api_overview(request):
    try:
        region = (request.GET.get("region") or "ALL").strip()
        product = (request.GET.get("product") or "ALL").strip()

        # MAIN TABLE (region x product_group)
        with connection.cursor() as cur:
            cur.execute("SET NAMES utf8mb4;")
            sql = """
                SELECT
                    COALESCE(m.original_value, 'Unknown') AS region_name,
                    COALESCE(p.product_group, 'Unknown') AS product_group,
                    SUM(s.sold_quantity) AS total_qty,
                    SUM(s.revenue)       AS total_rev,
                    SUM(s.cost_price)    AS total_cost,
                    SUM(s.profit)        AS total_profit
                FROM sales s
                LEFT JOIN product p ON s.product_id = p.product_id
                LEFT JOIN mapping m
                  ON m.encoded_value = s.distribution_channel_code_enc
                 AND m.variable_name = 'region'
                WHERE (%s = 'ALL' OR m.original_value = %s)
                  AND (%s = 'ALL' OR p.product_group = %s)
                GROUP BY COALESCE(m.original_value, 'Unknown'),
                         COALESCE(p.product_group, 'Unknown')
                HAVING SUM(s.sold_quantity) > 0
                ORDER BY total_rev DESC
                LIMIT 2000;
            """
            cur.execute(sql, [region, region, product, product])
            rows = cur.fetchall()

        result_rows = []
        for r in rows:
            region_name, product_group, qty, rev, cost, profit = r
            qty = _to_float(qty)
            rev = _to_float(rev)
            cost = _to_float(cost)
            profit = _to_float(profit)
            pct = round(100 * profit / rev, 2) if rev else 0
            result_rows.append({
                "region": region_name,
                "product": product_group,
                "quantity": qty,
                "revenue": rev,
                "cost": cost,
                "profit": profit,
                "profit_pct": pct,
            })

        # TIME SERIES (revenue + profit)
        with connection.cursor() as cur:
            sql_time = """
                SELECT year_num, month_num,
                       SUM(revenue) AS rev,
                       SUM(profit)  AS profit
                FROM sales s
                LEFT JOIN product p ON s.product_id = p.product_id
                WHERE (%s = 'ALL' OR p.product_group = %s)
                GROUP BY year_num, month_num
                ORDER BY year_num, month_num
            """
            cur.execute(sql_time, [product, product])
            trows = cur.fetchall()

        labels, rev_series, profit_series = [], [], []
        for yr, mo, rev, profit in trows:
            if yr is None or mo is None:
                continue
            labels.append(f"{int(yr)}-{int(mo):02d}")
            rev_series.append(_to_float(rev))
            profit_series.append(_to_float(profit))

        # ----------------------------------------------------------
        # NEW: CHART PRODUCT BY CATEGORY (not top 5)
        # ----------------------------------------------------------
        with connection.cursor() as cur:
            sql_cat = """
                SELECT COALESCE(p.product_group, 'Unknown') AS grp,
                       SUM(s.revenue) AS total_rev
                FROM sales s
                LEFT JOIN product p ON s.product_id = p.product_id
                WHERE (%s = 'ALL' OR p.product_group = %s)
                GROUP BY COALESCE(p.product_group, 'Unknown')
                ORDER BY total_rev DESC
            """
            cur.execute(sql_cat, [product, product])
            cat_rows = cur.fetchall()

        product_cat_labels = [r[0] for r in cat_rows]
        product_cat_values = [_to_float(r[1]) for r in cat_rows]

        # ----------------------------------------------------------
        # TOP 5 PRODUCTS (actual products)
        # ----------------------------------------------------------
        with connection.cursor() as cur:
            sql_top = """
                SELECT p.product_id, SUM(s.revenue) AS total_rev
                FROM sales s
                LEFT JOIN product p ON s.product_id = p.product_id
                WHERE p.product_id IS NOT NULL
                AND (%s = 'ALL' OR p.product_group = %s)
                GROUP BY p.product_id
                HAVING SUM(s.revenue) > 0
                ORDER BY total_rev DESC
            LIMIT 5;
            """
            cur.execute(sql_top, [product, product])
            top_rows = cur.fetchall()

        clean_rows = [(pid, _to_float(rev)) for pid, rev in top_rows if pid not in (None, "", "—")]

        # Nếu ít hơn 5 thì thôi, KHÔNG bù "—"
        top_labels = [r[0] for r in clean_rows]
        top_values = [r[1] for r in clean_rows]

        # ----------------------------------------------------------
        # KPIs
        # ----------------------------------------------------------
        total_rev = sum(x["revenue"] for x in result_rows) if result_rows else sum(rev_series)
        total_cost = sum(x["cost"] for x in result_rows) if result_rows else 0
        total_profit = sum(x["profit"] for x in result_rows) if result_rows else sum(profit_series)
        pct = round(100 * total_profit / total_rev, 2) if total_rev else 0

        kpis = {
            "revenue": round(total_rev, 2),
            "cost": round(total_cost, 2),
            "profit": round(total_profit, 2),
            "profit_pct": pct
        }

        return JsonResponse({
            "ok": True,
            "data": result_rows,
            "time": {
                "labels": labels,
                "revenue": rev_series,
                "profit": profit_series
            },
            "product_category": {
                "labels": product_cat_labels,
                "values": product_cat_values
            },
            "top5": {
                "labels": top_labels,
                "values": top_values
            },
            "kpis": kpis
        })

    except Exception as e:
        print("Lỗi API OVERVIEW:\n", traceback.format_exc())
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


# web/views.py

# ... (các hàm khác)

# =======================
# 4. OVERVIEW APIs
# =======================
@csrf_exempt
@login_required(login_url='login')
def download_overview(request):
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.chart import LineChart, Reference, BarChart
        from io import BytesIO
        import json
        from datetime import datetime

        payload = json.loads(request.body.decode("utf-8")) if request.body else {}

        # ===============================
        # LẤY KPI TỪ PAYLOAD (UI) – KHÔNG TÍNH LẠI
        # ===============================
        kpis = payload.get("kpis", {}) or {}

        total_rev = float(kpis.get("revenue") or 0)
        total_profit = float(kpis.get("profit") or 0)
        total_cost = float(kpis.get("cost") or (total_rev - total_profit))
        pct = float(kpis.get("profit_pct") or 0)

        # ===============================
        # TIME SERIES (UI gửi lên rồi)
        # ===============================
        time_obj = payload.get("time", {}) or {}
        labels = time_obj.get("labels", []) or []
        revs = time_obj.get("revenue", []) or []
        profs = time_obj.get("profit", []) or []

        # ===============================
        # TOP 5 PRODUCTS (UI)
        # ===============================
        top5 = payload.get("top5", {}) or {}
        top_labels = top5.get("labels", []) or []
        top_values = top5.get("values", []) or []

        # ===============================
        # TẠO EXCEL
        # ===============================
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tổng Quan"

        bold = Font(bold=True, size=14)

        # Header
        timestamp_vi = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A1"] = f"BÁO CÁO TỔNG QUAN — Xuất lúc {timestamp_vi}"
        ws["A1"].font = bold
        ws.merge_cells("A1:E1")

        # KPIs
        ws.append([])
        ws.append(["KPIs"])
        ws["A3"].font = bold

        ws.append(["Doanh thu", total_rev])
        ws.append(["Chi phí", total_cost])
        ws.append(["Lợi nhuận", total_profit])
        ws.append(["Tỷ suất LN (%)", pct])

        # TIME SERIES TABLE
        ws.append([])
        ws.append(["DỮ LIỆU THEO THÁNG"])
        ws["A{}".format(ws.max_row)].font = bold
        ws.append(["Thời điểm", "Doanh thu", "Lợi nhuận"])

        for ym, rev, prof in zip(labels, revs, profs):
            ws.append([ym, rev, prof])

        # ===============================
        # LINE CHART (Revenue & Profit)
        # ===============================
        chart = LineChart()
        chart.title = "Doanh thu & Lợi nhuận theo tháng"
        chart.style = 10
        chart.y_axis.title = "Giá trị"
        chart.x_axis.title = "Thời gian"

        data_start = ws.max_row - len(labels)

        data = Reference(ws, min_col=2, min_row=data_start, max_col=3, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "G5")

        # ===============================
        # TOP 5 TABLE
        # ===============================
        ws.append([])
        ws.append(["TOP 5 SẢN PHẨM"])
        ws["A{}".format(ws.max_row)].font = bold

        ws.append(["Sản phẩm", "Doanh thu"])

        for lab, val in zip(top_labels, top_values):
            ws.append([lab, val])

        # ===============================
        # TOP 5 BAR CHART
        # ===============================
        chart2 = BarChart()
        chart2.title = "Top 5 sản phẩm theo doanh thu"
        chart2.y_axis.title = "Doanh thu"
        chart2.x_axis.title = "Sản phẩm"

        top_start = ws.max_row - len(top_labels)

        data2 = Reference(ws, min_col=2, min_row=top_start, max_row=ws.max_row)
        cats2 = Reference(ws, min_col=1, min_row=top_start+1, max_row=ws.max_row)

        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4

        ws.add_chart(chart2, "G25")

        # ===============================
        # EXPORT FILE
        # ===============================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Bao_cao_tong_quan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return response

    except Exception as e:
        print("❌ Lỗi download_overview:", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


# HÀM XUẤT FILE DỰ BÁO (cho /api/forecast/export/)
@csrf_exempt
@login_required(login_url='login')
@require_http_methods(["POST"])
def export_forecast(request):
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        product_id = payload.get("product_id", "ALL")
        horizon = int(payload.get("horizon", 3))
        model_name = payload.get("model", "baseline").lower()

        # 1. Lấy dữ liệu lịch sử
        data_hist = _fetch_sales_weekly(product_id, aggregate_by_month=True)
        actual = data_hist["actual"]
        months_hist = data_hist["months"]

        if not actual:
            return JsonResponse({"ok": False, "error": "Không có dữ liệu lịch sử."}, status=400)

        # 2. Chạy mô hình (baseline)
        fitted, forecast = _sma_forecast(actual, horizon=horizon)

        # 3. Tạo file Excel bằng openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dự báo nhu cầu"

        ws.append(["Loại", "Tháng", "Sản lượng"])

        for i, q in enumerate(actual):
            month_label = f"T{months_hist[i]}" if i < len(months_hist) else f"T{i+1}"
            ws.append(["Thực tế", month_label, round(q, 2)])

        last_hist = months_hist[-1] if months_hist else 0
        for i, q in enumerate(forecast):
            month_label = f"T{last_hist + i + 1}"
            ws.append(["Dự báo", month_label, round(q, 2)])

        filename = f"Du_bao_Nhu_cau_{product_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        response = HttpResponse(
            file_stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # 5. Lưu lịch sử y như các loại báo cáo khác
        try:
            file_size_kb = round(len(file_stream.getvalue()) / 1024, 2)
            params_data = {
                "product_id": product_id,
                "region": payload.get("region", "ALL"),
                "model": model_name,
                "horizon": horizon,
            }
            ReportDownloadHistory.objects.create(
                user=request.user,
                payload={
                    "type": "Dự báo nhu cầu",
                    "params": params_data,
                },
                filename=filename,
                file_size_kb=file_size_kb,
            )
            print("[Forecast] Saved history OK")

        except Exception as e:
            print("❌ Lỗi lưu lịch sử forecast:", e)

        return response

    except Exception as e:
        print("❌ Forecast Export Error:", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

@csrf_exempt
@login_required(login_url='login')
def download_report(request):
    try:
        import json
        from datetime import datetime
        from io import BytesIO
        import openpyxl

        payload = json.loads(request.body.decode('utf-8')) if request.body else {}

        # tạo workbook
        wb = openpyxl.Workbook()

        # gọi hàm build nội dung
        create_excel_content(payload, wb)

        # output vào buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # tên file chuẩn (YYYYMMDD_HHMMSS)
        filename = f"Bao_cao_gia_khuyen_mai_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        print("❌ Lỗi download_report:", e)
        return JsonResponse({"ok": False, "error": str(e)})

